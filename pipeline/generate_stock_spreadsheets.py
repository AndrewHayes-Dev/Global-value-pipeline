"""
generate_stock_spreadsheets.py — Global Value Pipeline
Generates one Excel workbook per index, uploaded to R2 at spreadsheets/{index_id}.xlsx
Includes an 'Industry' column not present in the original pipeline.
"""

import io
import math
import sys
from datetime import datetime, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit('ERROR: pip install openpyxl')

INDEX_NAMES = {
    'sp500':       'S&P 500',
    'djia':        'DJIA',
    'nasdaq':      'NASDAQ-100',
    'russell2000': 'Russell 2000',
    'xao':         'ASX All Ordinaries',
}

COLUMNS = [
    ('rank_in_sector',       'Rank',                 'id',    'int',     6),
    ('ticker',               'Ticker',               'id',    'str',     9),
    ('company_name',         'Company Name',         'id',    'str',    28),
    ('industry',             'Industry',             'id',    'str',    24),
    ('score',                'Buffett Score',        'id',    'f1',     12),
    ('current_price',        'Price ($)',            'yahoo', '$',      11),
    ('pe_ratio',             'P/E Ratio',            'mixed', 'f2',     10),
    ('pb_ratio',             'P/B Ratio',            'mixed', 'f2',     10),
    ('dividend_yield',       'Dividend Yield',       'mixed', 'pct',    14),
    ('revenue',              'Revenue',              'yahoo', '$bm',    14),
    ('net_income',           'Net Income',           'yahoo', '$bm',    14),
    ('free_cash_flow',       'Free Cash Flow',       'mixed', '$bm',    14),
    ('debt_equity',          'Debt / Equity',        'mixed', 'f2',     12),
    ('intrinsic_value',      'Intrinsic Value',      'yahoo', '$',      14),
    ('margin_of_safety',     'Margin of Safety',     'yahoo', 'pct_raw',16),
    ('owner_earnings',       'Owner Earnings',       'yahoo', '$bm',    14),
    ('roic',                 'ROIC',                 'mixed', 'pct',    10),
    ('roe',                  'ROE',                  'yahoo', 'pct',    10),
    ('peg_ratio',            'PEG Ratio',            'yahoo', 'f2',     10),
    ('interest_coverage',    'Interest Coverage',    'yahoo', 'f1',     16),
    ('gross_margin',         'Gross Margin',         'yahoo', 'pct',    13),
    ('current_ratio',        'Current Ratio',        'yahoo', 'f2',     13),
    ('eps_growth',           'EPS Growth',           'yahoo', 'pct',    11),
    ('roe_3yr_avg',          'ROE 3yr Avg',          'yahoo', 'pct',    12),
    ('earnings_consistency', 'Earnings Consistency', 'yahoo', 'int4',   18),
    ('book_value_growth',    'Book Value Growth',    'yahoo', 'pct',    16),
    ('net_net_ratio',        'Net-Net Ratio',        'yahoo', 'f2',     13),
    ('revenue_growth',       'Revenue Growth',       'yahoo', 'pct',    14),
]

_C_ID    = 'D9D9D9'
_C_YAHOO = 'BDD7EE'
_C_MIXED = 'FCE4D6'
_C_TITLE = '1F3864'
_C_EVEN  = 'F5FAFF'
_C_ODD   = 'FFFFFF'

_HDR_FG = {'id': '404040', 'yahoo': '1F3864', 'mixed': '843C00'}
_HDR_BG = {'id': _C_ID, 'yahoo': _C_YAHOO, 'mixed': _C_MIXED}

_thin   = Side(style='thin',   color='CCCCCC')
_med    = Side(style='medium', color='1F3864')
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_TITLE_BORDER = Border(left=_med, right=_med, top=_med, bottom=_med)


def _fill(c):
    return PatternFill('solid', fgColor=c)


def _fmt(raw, fmt):
    if raw is None:
        return None
    if fmt == 'str':
        return str(raw) if raw is not None else None
    try:
        n = float(raw)
    except (TypeError, ValueError):
        return str(raw)
    if math.isnan(n) or math.isinf(n):
        return None
    if fmt == 'int':
        return int(round(n))
    if fmt == 'int4':
        return f'{int(round(n))}/4'
    if fmt == 'f1':
        return round(n, 1)
    if fmt == 'f2':
        return round(n, 2)
    if fmt == '$':
        return f'${n:,.2f}'
    if fmt == '$bm':
        sign = '-' if n < 0 else ''
        a = abs(n)
        if a >= 1e12: return f'{sign}${a / 1e12:.2f}T'
        if a >= 1e9:  return f'{sign}${a / 1e9:.2f}B'
        if a >= 1e6:  return f'{sign}${a / 1e6:.2f}M'
        if a >= 1e3:  return f'{sign}${a / 1e3:.1f}K'
        return f'{sign}${a:.0f}'
    if fmt == 'pct':
        return f'{n * 100:.1f}%'
    if fmt == 'pct_raw':
        return f'{n:.1f}%'
    return n


def _title_cell(ws, value, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value=value)
    c.fill = _fill(_C_TITLE)
    c.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = _TITLE_BORDER
    return c


def _header_cell(ws, row, col, label, source):
    c = ws.cell(row=row, column=col, value=label)
    c.fill = _fill(_HDR_BG[source])
    c.font = Font(name='Calibri', size=9, bold=True, color=_HDR_FG[source])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = _BORDER
    return c


def _data_cell(ws, row, col, value, is_even):
    halign = 'left' if isinstance(value, str) else 'right'
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(_C_EVEN if is_even else _C_ODD)
    c.font = Font(name='Calibri', size=9, color='000000')
    c.alignment = Alignment(horizontal=halign, vertical='center')
    c.border = _BORDER
    return c


def _build_sector_sheet(wb, sector, index_name, generated_at):
    tab_name = sector['name'][:31]
    ws = wb.create_sheet(title=tab_name)
    stocks = sorted(sector.get('stocks', []), key=lambda s: s.get('rank_in_sector', 9999))
    num_cols = len(COLUMNS)

    title_text = (
        f'{index_name} — {sector["name"]}     '
        f'■ Grey = Identifier / Score     '
        f'■ Blue = Yahoo Finance     '
        f'■ Orange = Yahoo Finance (primary) / FMP (fallback)     '
        f'Updated: {generated_at}'
    )
    _title_cell(ws, title_text, num_cols)
    ws.row_dimensions[1].height = 20

    for ci, (_, label, source, _, _) in enumerate(COLUMNS, start=1):
        _header_cell(ws, 2, ci, label, source)
    ws.row_dimensions[2].height = 30

    for ci, (*_, col_w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w

    for ri, stock in enumerate(stocks):
        row = ri + 3
        is_even = ri % 2 == 0
        for ci, (key, _, _, fmt, _) in enumerate(COLUMNS, start=1):
            raw = stock.get(key)
            _data_cell(ws, row, ci, _fmt(raw, fmt), is_even)
        ws.row_dimensions[row].height = 15

    ws.freeze_panes = ws.cell(row=3, column=6)
    last_data_row = 2 + len(stocks) if stocks else 2
    ws.auto_filter.ref = f'A2:{get_column_letter(num_cols)}{last_data_row}'


def _build_workbook(index_data, index_name, generated_at):
    wb = Workbook()
    wb.remove(wb.active)
    for sector in index_data.get('sectors', []):
        if sector.get('stocks'):
            _build_sector_sheet(wb, sector, index_name, generated_at)
    return wb


def generate_all(all_index_data: dict, bucket: str) -> None:
    from uploader import upload_bytes
    generated_at = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    print('\nGenerating stock data spreadsheets...')
    for index_id, index_data in all_index_data.items():
        index_name = INDEX_NAMES.get(index_id, index_id.upper())
        print(f'  {index_name:20s}', end='', flush=True)
        wb = _build_workbook(index_data, index_name, generated_at)
        if not wb.worksheets:
            print('(no data — skipped)')
            continue
        buf = io.BytesIO()
        wb.save(buf)
        upload_bytes(
            f'spreadsheets/{index_id}.xlsx',
            buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            bucket,
        )
